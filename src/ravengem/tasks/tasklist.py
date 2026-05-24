"""Parse a metabolic task list (port of RAVEN ``parseTaskList``).

A task list defines, per task, allowed inputs/outputs, optional extra reactions
(equations), reaction-bound changes, and whether the task *should fail*. Tasks
are checked with :func:`ravengem.tasks.check_tasks`.

The file is tab-delimited (``.txt``/``.tsv``) or Excel (``.xlsx``, sheet ``TASKS``;
needs the ``[excel]`` extra). Recognised columns (the only required one is ``ID``):

    ID · DESCRIPTION · IN · IN LB · IN UB · OUT · OUT LB · OUT UB ·
    EQU · EQU LB · EQU UB · CHANGED RXN · CHANGED LB · CHANGED UB ·
    SHOULD FAIL · PRINT FLUX · COMMENTS

A task spans consecutive rows; only its first row carries an ID. Metabolites are
written ``name[compartment]``; several in one cell are separated by ``;`` (sharing
that row's bounds). ``IN``/``OUT`` default LB 0, UB 1000; ``EQU`` defaults LB
-1000 if reversible (``<=>``) else 0, UB 1000. The special tokens ``ALLMETS`` and
``ALLMETSIN[comp]`` allow free uptake/excretion of all metabolites (only the upper
bound is used).
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

_COLUMNS = (
    "ID", "DESCRIPTION", "IN", "IN LB", "IN UB", "OUT", "OUT LB", "OUT UB",
    "EQU", "EQU LB", "EQU UB", "CHANGED RXN", "CHANGED LB", "CHANGED UB",
    "SHOULD FAIL", "PRINT FLUX", "COMMENTS",
)


@dataclass
class Task:
    """One metabolic task. Bounds are ``(metabolite_or_reaction, lb, ub)`` triples."""

    id: str
    description: str = ""
    should_fail: bool = False
    print_fluxes: bool = False
    comments: str = ""
    inputs: list[tuple[str, float, float]] = field(default_factory=list)
    outputs: list[tuple[str, float, float]] = field(default_factory=list)
    equations: list[tuple[str, float, float]] = field(default_factory=list)
    changed: list[tuple[str, float, float]] = field(default_factory=list)


def _truthy(value: str) -> bool:
    return value.strip().lower() not in ("", "0", "false", "no")


def _num(value: str, default: float) -> float:
    value = value.strip()
    return float(value) if value else default


def _read_rows(path: str | Path) -> list[list[str]]:
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - optional dep
            raise ImportError("Reading .xlsx task lists needs the '[excel]' extra (openpyxl).") from exc
        ws = load_workbook(path, data_only=True)["TASKS"]
        return [["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    with open(path, encoding="utf-8", newline="") as handle:
        return [row for row in csv.reader(handle, delimiter="\t")]


def parse_task_list(path: str | Path) -> list[Task]:
    """Parse a task-list file into :class:`Task` objects."""
    rows = _read_rows(path)
    header_idx = next(
        (i for i, r in enumerate(rows) if any(c.strip().upper() == "ID" for c in r)), None
    )
    if header_idx is None:
        raise ValueError(f"{path}: no header row with an 'ID' column found.")
    header = [c.strip().upper() for c in rows[header_idx]]
    col = {name: header.index(name) for name in _COLUMNS if name in header}

    def cell(row: list[str], name: str) -> str:
        i = col.get(name)
        return row[i].strip() if i is not None and i < len(row) else ""

    tasks: list[Task] = []
    current: Task | None = None
    for row in rows[header_idx + 1:]:
        if not any(c.strip() for c in row):
            continue
        rid = cell(row, "ID")
        if rid.startswith("#"):
            continue
        if rid:
            current = Task(
                id=rid,
                description=cell(row, "DESCRIPTION"),
                should_fail=_truthy(cell(row, "SHOULD FAIL")),
                print_fluxes=_truthy(cell(row, "PRINT FLUX")),
                comments=cell(row, "COMMENTS"),
            )
            tasks.append(current)
        if current is None:
            continue
        _add_row(current, row, cell)
    return tasks


def _add_row(task: Task, row: list[str], cell) -> None:
    if inp := cell(row, "IN"):
        lb, ub = _num(cell(row, "IN LB"), 0.0), _num(cell(row, "IN UB"), 1000.0)
        task.inputs += [(m.strip(), lb, ub) for m in inp.split(";") if m.strip()]
    if out := cell(row, "OUT"):
        lb, ub = _num(cell(row, "OUT LB"), 0.0), _num(cell(row, "OUT UB"), 1000.0)
        task.outputs += [(m.strip(), lb, ub) for m in out.split(";") if m.strip()]
    if equ := cell(row, "EQU"):
        lb = _num(cell(row, "EQU LB"), -1000.0 if "<=>" in equ else 0.0)
        ub = _num(cell(row, "EQU UB"), 1000.0)
        task.equations.append((equ.strip(), lb, ub))
    if chg := cell(row, "CHANGED RXN"):
        lb, ub = _num(cell(row, "CHANGED LB"), -1000.0), _num(cell(row, "CHANGED UB"), 1000.0)
        task.changed += [(r.strip(), lb, ub) for r in chg.split(";") if r.strip()]
