"""Export a model to the RAVEN Microsoft Excel format.

Writes the five-sheet RAVEN xlsx layout — RXNS, METS, COMPS, GENES, MODEL — pulling
RAVEN-specific values back out of cobra's ``annotation`` / ``notes`` (where the
raven_python YAML reader stashes them). Excel *import* is intentionally not provided.

Requires the optional ``openpyxl`` dependency (``pip install raven_python[excel]``).
"""
from __future__ import annotations

from pathlib import Path

import cobra

from raven_python.utils.parse import subsystem_to_str


def _miriam_string(annotation: dict, exclude: tuple[str, ...] = ()) -> str:
    """RAVEN MIRIAM column: ``namespace/id;namespace/id2;...`` (sorted)."""
    parts: list[str] = []
    for namespace in sorted(annotation):
        if namespace in exclude:
            continue
        values = annotation[namespace]
        if isinstance(values, str):
            values = [values]
        parts.extend(f"{namespace}/{value}" for value in values)
    return ";".join(parts)


def _equation(rxn: cobra.Reaction) -> str:
    """Human-readable equation in RAVEN ``name[comp]`` form."""

    def side(items):
        return " + ".join(
            f"{abs(coef):g} {met.name}[{met.compartment}]" for met, coef in items
        )

    reactants = [(m, c) for m, c in rxn.metabolites.items() if c < 0]
    products = [(m, c) for m, c in rxn.metabolites.items() if c > 0]
    arrow = " <=> " if rxn.reversibility else " => "
    return f"{side(reactants)}{arrow}{side(products)}"


def _ec_codes(rxn: cobra.Reaction) -> str:
    codes = rxn.annotation.get("ec-code", [])
    if isinstance(codes, str):
        codes = [codes]
    return ";".join(codes)


def export_to_excel(
    model: cobra.Model, path: str | Path, *, sort_ids: bool = False
) -> None:
    """Write ``model`` to a RAVEN-format ``.xlsx`` file.

    Parameters
    ----------
    sort_ids
        If True, write reactions/metabolites/genes sorted alphabetically by ID
        (the model itself is not modified).
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise ImportError(
            "export_to_excel requires openpyxl. Install it with "
            "`pip install raven_python[excel]` (or `pip install openpyxl`)."
        ) from exc

    reactions = sorted(model.reactions, key=lambda r: r.id) if sort_ids else list(model.reactions)
    metabolites = (
        sorted(model.metabolites, key=lambda m: m.id) if sort_ids else list(model.metabolites)
    )
    genes = sorted(model.genes, key=lambda g: g.id) if sort_ids else list(model.genes)
    metadata = dict(model.notes.get("metaData", {})) if model.notes else {}

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    # --- RXNS ---
    ws = wb.create_sheet("RXNS")
    ws.append(
        ["#", "ID", "NAME", "EQUATION", "EC-NUMBER", "GENE ASSOCIATION", "LOWER BOUND",
         "UPPER BOUND", "OBJECTIVE", "COMPARTMENT", "MIRIAM", "SUBSYSTEM",
         "REPLACEMENT ID", "NOTE", "REFERENCE", "CONFIDENCE SCORE"]
    )
    for r in reactions:
        subsystem = subsystem_to_str(r.subsystem)
        ws.append([
            None, r.id, r.name, _equation(r), _ec_codes(r), r.gene_reaction_rule,
            r.lower_bound, r.upper_bound,
            r.objective_coefficient or None, None,
            _miriam_string(r.annotation, exclude=("ec-code",)), subsystem, None,
            r.notes.get("note"), r.notes.get("references"), r.notes.get("confidence_score"),
        ])

    # --- METS ---
    ws = wb.create_sheet("METS")
    ws.append(["#", "ID", "NAME", "UNCONSTRAINED", "MIRIAM", "COMPOSITION", "InChI",
               "COMPARTMENT", "REPLACEMENT ID", "CHARGE"])
    for m in metabolites:
        inchi = m.notes.get("inchis")
        ws.append([
            None, f"{m.name}[{m.compartment}]", m.name, None,
            _miriam_string(m.annotation, exclude=("smiles",)),
            m.formula, inchi, m.compartment, m.id, m.charge,
        ])

    # --- COMPS ---
    ws = wb.create_sheet("COMPS")
    ws.append(["#", "ABBREVIATION", "NAME", "INSIDE", "MIRIAM"])
    comps = sorted(model.compartments) if sort_ids else list(model.compartments)
    for cid in comps:
        ws.append([None, cid, model.compartments.get(cid, ""), None, None])

    # --- GENES ---
    if genes:
        ws = wb.create_sheet("GENES")
        ws.append(["#", "NAME", "MIRIAM", "SHORT NAME", "COMPARTMENT"])
        for g in genes:
            ws.append([None, g.id, _miriam_string(g.annotation), g.name, None])

    # --- MODEL ---
    ws = wb.create_sheet("MODEL")
    ws.append(["#", "ID", "NAME", "TAXONOMY", "DEFAULT LOWER", "DEFAULT UPPER",
               "CONTACT GIVEN NAME", "CONTACT FAMILY NAME", "CONTACT EMAIL",
               "ORGANIZATION", "NOTES"])
    ws.append([
        None, model.id or "blankID", model.name or "blankName",
        metadata.get("taxonomy"), metadata.get("defaultLB"), metadata.get("defaultUB"),
        metadata.get("givenName"), metadata.get("familyName"), metadata.get("email"),
        metadata.get("organization"), metadata.get("note"),
    ])

    wb.save(str(path))
