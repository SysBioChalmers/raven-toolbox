"""Export a model to the RAVEN Microsoft Excel format.

Writes the five-sheet RAVEN xlsx layout — RXNS, METS, COMPS, GENES, MODEL — pulling
RAVEN-specific values back out of cobra's ``annotation`` / ``notes`` (where the
raven_toolbox YAML reader stashes them). For enzyme-constrained (GECKO) models that
carry a populated ``model.ec`` substructure, two further sheets are added, ENZYMES
and ENZRXNS, holding the ec data. Excel *import* is intentionally not provided.

Requires the optional ``openpyxl`` dependency (``pip install raven_toolbox[excel]``).
"""
from __future__ import annotations

import math
from pathlib import Path

import cobra

from raven_toolbox.utils.parse import subsystem_to_str


def _miriam_string(annotation: dict, exclude: tuple[str, ...] = ()) -> str:
    """RAVEN MIRIAM column: ``namespace/id;namespace/id2;...`` (sorted)."""
    parts: list[str] = []
    for namespace in sorted(annotation):
        if namespace in exclude:
            continue
        values = annotation[namespace]
        if isinstance(values, str):
            values = [values]
        parts.extend(f"{namespace}/{value}" for value in values)
    return ";".join(parts)


def _equation(rxn: cobra.Reaction) -> str:
    """Human-readable equation in RAVEN ``name[comp]`` form."""

    def side(items):
        return " + ".join(
            f"{abs(coef):g} {met.name}[{met.compartment}]" for met, coef in items
        )

    reactants = [(m, c) for m, c in rxn.metabolites.items() if c < 0]
    products = [(m, c) for m, c in rxn.metabolites.items() if c > 0]
    arrow = " <=> " if rxn.reversibility else " => "
    return f"{side(reactants)}{arrow}{side(products)}"


def _ec_codes(rxn: cobra.Reaction) -> str:
    codes = rxn.annotation.get("ec-code", [])
    if isinstance(codes, str):
        codes = [codes]
    return ";".join(codes)


def _blank_if_nan(value: float) -> float | None:
    """Numeric cell value, writing NaN (the EcData "unknown" sentinel) as blank."""
    value = float(value)
    return None if math.isnan(value) else value


def _blank_if_empty(value: str) -> str | None:
    """String cell value, writing an empty string as a blank cell."""
    return value or None


def _fmt_count(value: float) -> str:
    """Subunit count as an integer string when integral, else as-is."""
    value = float(value)
    return str(int(value)) if value.is_integer() else f"{value:g}"


def _enzyme_pairs(ec, coupling, i: int) -> str | None:
    """ENZRXNS 'ENZYMES' column: ``enzyme:count;...`` subunit stoichiometry for
    ec-reaction ``i``, read from the ``rxn_enz_mat`` coupling row. Blank when
    the reaction has no associated enzymes."""
    if i >= coupling.shape[0]:
        return None
    row = coupling.getrow(i)
    if row.nnz == 0:
        return None
    pairs = sorted(zip(row.indices.tolist(), row.data.tolist(), strict=True))
    return ";".join(f"{ec.enzymes[j]}:{_fmt_count(v)}" for j, v in pairs)


def export_to_excel(
    model: cobra.Model, path: str | Path, *, sort_ids: bool = False
) -> None:
    """Write ``model`` to a RAVEN-format ``.xlsx`` file.

    For enzyme-constrained models (a populated :class:`~raven_toolbox.io.EcData`
    on ``model.ec``), two further export-only sheets are written: ENZYMES (one
    row per enzyme) and ENZRXNS (one row per ec-reaction).

    Parameters
    ----------
    sort_ids
        If True, write reactions/metabolites/genes sorted alphabetically by ID
        (the model itself is not modified). The ec sheets are not reordered.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise ImportError(
            "export_to_excel requires openpyxl. Install it with "
            "`pip install raven_toolbox[excel]` (or `pip install openpyxl`)."
        ) from exc

    reactions = sorted(model.reactions, key=lambda r: r.id) if sort_ids else list(model.reactions)
    metabolites = (
        sorted(model.metabolites, key=lambda m: m.id) if sort_ids else list(model.metabolites)
    )
    genes = sorted(model.genes, key=lambda g: g.id) if sort_ids else list(model.genes)
    metadata = dict(model.notes.get("metaData", {})) if model.notes else {}

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    # --- RXNS ---
    ws = wb.create_sheet("RXNS")
    ws.append(
        ["#", "ID", "NAME", "EQUATION", "EC-NUMBER", "GENE ASSOCIATION", "LOWER BOUND",
         "UPPER BOUND", "OBJECTIVE", "COMPARTMENT", "MIRIAM", "SUBSYSTEM",
         "REPLACEMENT ID", "NOTE", "REFERENCE", "CONFIDENCE SCORE"]
    )
    for r in reactions:
        subsystem = subsystem_to_str(r.subsystem)
        ws.append([
            None, r.id, r.name, _equation(r), _ec_codes(r), r.gene_reaction_rule,
            r.lower_bound, r.upper_bound,
            r.objective_coefficient or None, None,
            _miriam_string(r.annotation, exclude=("ec-code",)), subsystem, None,
            r.notes.get("note"), r.notes.get("references"), r.notes.get("confidence_score"),
        ])

    # --- METS ---
    ws = wb.create_sheet("METS")
    ws.append(["#", "ID", "NAME", "UNCONSTRAINED", "MIRIAM", "COMPOSITION", "InChI",
               "COMPARTMENT", "REPLACEMENT ID", "CHARGE"])
    for m in metabolites:
        inchi = m.notes.get("inchis")
        ws.append([
            None, f"{m.name}[{m.compartment}]", m.name, None,
            _miriam_string(m.annotation, exclude=("smiles",)),
            m.formula, inchi, m.compartment, m.id, m.charge,
        ])

    # --- COMPS ---
    ws = wb.create_sheet("COMPS")
    ws.append(["#", "ABBREVIATION", "NAME", "INSIDE", "MIRIAM"])
    comps = sorted(model.compartments) if sort_ids else list(model.compartments)
    for cid in comps:
        ws.append([None, cid, model.compartments.get(cid, ""), None, None])

    # --- GENES ---
    if genes:
        ws = wb.create_sheet("GENES")
        ws.append(["#", "NAME", "MIRIAM", "SHORT NAME", "COMPARTMENT"])
        for g in genes:
            ws.append([None, g.id, _miriam_string(g.annotation), g.name, None])

    # --- MODEL ---
    ws = wb.create_sheet("MODEL")
    ws.append(["#", "ID", "NAME", "TAXONOMY", "DEFAULT LOWER", "DEFAULT UPPER",
               "CONTACT GIVEN NAME", "CONTACT FAMILY NAME", "CONTACT EMAIL",
               "ORGANIZATION", "NOTES"])
    ws.append([
        None, model.id or "blankID", model.name or "blankName",
        metadata.get("taxonomy"), metadata.get("defaultLB"), metadata.get("defaultUB"),
        metadata.get("givenName"), metadata.get("familyName"), metadata.get("email"),
        metadata.get("organization"), metadata.get("note"),
    ])

    # --- ENZYMES / ENZRXNS (enzyme-constrained GECKO models) ---
    # When the model carries a populated ec substructure (model.ec), write its
    # contents to two export-only sheets, mirroring RAVEN's exportToExcelFormat.
    # The YAML format remains the round-trippable format for ecModels; the
    # enzyme-reaction coupling (ec.rxn_enz_mat) is written in readable form as
    # the 'enzyme:count' ENZYMES column of the ENZRXNS sheet.
    ec = getattr(model, "ec", None)
    if ec is not None and (ec.n_enzymes or ec.n_rxns):
        ws = wb.create_sheet("ENZYMES")
        ws.append(["#", "ID", "GENE", "MW", "SEQUENCE", "CONC"])
        for i in range(ec.n_enzymes):
            ws.append([
                None, ec.enzymes[i], ec.genes[i], _blank_if_nan(ec.mw[i]),
                _blank_if_empty(ec.sequence[i]), _blank_if_nan(ec.concs[i]),
            ])
        # MW (column D) shown without decimals, CONC (column F) with 5 decimals
        for r in range(2, ec.n_enzymes + 2):
            ws.cell(row=r, column=4).number_format = "0"
            ws.cell(row=r, column=6).number_format = "0.00000"

        ws = wb.create_sheet("ENZRXNS")
        ws.append(["#", "ID", "KCAT", "SOURCE", "NOTE", "EC-NUMBER", "ENZYMES"])
        coupling = ec.rxn_enz_mat.tocsr()
        for i in range(ec.n_rxns):
            ws.append([
                None, ec.rxns[i], _blank_if_nan(ec.kcat[i]),
                _blank_if_empty(ec.source[i]), _blank_if_empty(ec.notes[i]),
                _blank_if_empty(ec.eccodes[i]), _enzyme_pairs(ec, coupling, i),
            ])

    wb.save(str(path))
