"""Tests for raven_python.utils.parse helpers."""
from __future__ import annotations

import cobra

from raven_python.utils.parse import parse_name_comp, subsystem_to_str


def test_parse_name_comp_basic():
    assert parse_name_comp("ATP[c]") == ("ATP", "c")
    assert parse_name_comp("ATP") == ("ATP", None)
    assert parse_name_comp("weird[name][m]") == ("weird[name]", "m")


def test_subsystem_to_str_scalar_and_none():
    assert subsystem_to_str("glycolysis") == "glycolysis"
    assert subsystem_to_str("") == ""
    assert subsystem_to_str(None) == ""


def test_subsystem_to_str_joins_list_without_data_loss():
    # A multi-subsystem reaction keeps every part (unlike taking only the first).
    assert subsystem_to_str(["glycolysis", "TCA cycle"]) == "glycolysis;TCA cycle"
    # Empty parts are dropped; the rest survive.
    assert subsystem_to_str(["", "x"]) == "x"


def test_subsystem_to_str_coerces_non_string_items():
    # Non-string items must not crash (the old excel.py ";".join did).
    assert subsystem_to_str([1, "name"]) == "1;name"


def test_excel_export_handles_list_subsystem(tmp_path):
    import pytest

    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from raven_python.io.excel import export_to_excel

    model = cobra.Model("m")
    a = cobra.Metabolite("a_c", compartment="c")
    b = cobra.Metabolite("b_c", compartment="c")
    model.add_metabolites([a, b])
    r = cobra.Reaction("R1", lower_bound=-1000, upper_bound=1000)
    model.add_reactions([r])
    r.add_metabolites({a: -1, b: 1})
    r.subsystem = ["glycolysis", "TCA cycle"]  # list, would crash old ";".join

    out = tmp_path / "m.xlsx"
    export_to_excel(model, out)
    wb = load_workbook(out)
    ws = wb["RXNS"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    sub_col = header.index("SUBSYSTEM")
    row1 = [c.value for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
    assert row1[sub_col] == "glycolysis;TCA cycle"
